import os
import requests
from datetime import datetime
from zoneinfo import ZoneInfo
import pyodbc
import pandas as pd
from azure.storage.blob import BlobServiceClient
import io
from openpyxl import load_workbook
from openpyxl import Workbook  # load_workbookの返り値bookの型を表現するときにのみ使う
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import TableColumn
from openpyxl.styles import Alignment
import re
import zipfile
import azure.functions as func
import logging


# アプリケーションの初期化
app = func.FunctionApp()


def get_env_or_raise(key: str) -> str:
    """Azureの環境変数から指定されたキーの値を返す。存在しなかったら例外を投げる。"""

    value = os.getenv(key)
    if not value:
        raise EnvironmentError(f"環境変数 {key} が設定されていません。Azure Functions の Application Settings に追加してください。")
    return value


def download_sql_text(sql_url: str) -> list[str]:
    """指定URLからSQLテキストをダウンロードして返す。"""

    # timeoutはblobからsqlファイルをダウンロード完了するまでの制限時間
    response = requests.get(sql_url, timeout=600)
    response.raise_for_status()
    response.encoding = response.encoding or 'utf-8'
    return response.text


def execute_sql_to_df(
    conn_str: str,
    sql_text: str
) -> pd.DataFrame:
    """pyodbc経由でSQLを実行してpandas.DataFrameを返す。"""

    # timeoutはクエリ実行から完了までの制限時間
    # autocommitはfalseでok（SELECTのみであればsqlのcommit()不要なので）
    with pyodbc.connect(conn_str, autocommit=False, timeout=600) as conn:
        df = pd.read_sql(sql_text, conn)
        df = df.fillna('NULL')
    
    return df


def replace_null(
    df: pd.DataFrame,
    names_df_list: list[pd.DataFrame],
    names_df_order: int,
    code_column: str,
    name_column: str
) -> None:
    """1と2の企業名/社員名 のNULLになっている箇所を正確な名前に置き換える"""

    for n in range(len(df)):
        if df.at[n, name_column] == 'NULL':
            code_that_have_null_name = df.at[n, code_column]
            names_df = names_df_list[names_df_order]

            # bool_series_whether_match_code は [False, False, True, False] のような pd.Series
            bool_series_whether_match_code = names_df[code_column] == code_that_have_null_name
            del_indexes = names_df.index[bool_series_whether_match_code]
            
            if len(del_indexes) == 0:
                logging.info(f'NULLの名前を持つ{code_column} {code_that_have_null_name} が名前取得用のdfには存在しなかったため、dfの該当セルがある行は削除します。')
                # 1で削除された社員は物理削除になる仕様で正しい接続先DBにも存在しなくなっているので、集計対象外としてレコードごと削除
                df.drop(n, inplace=True)
            else:
                # DBの仕様上1つしかないのが確定しているので[0]と断定してOK
                del_index = del_indexes[0]
                correct_name = names_df.at[del_index, name_column]
                df.at[n, name_column] = correct_name


def apply_column_style(
    ws: Worksheet,
    left_align_cols: list[str]
) -> None:
    """
    【追加】指定されたシートのデータ行(4行目以降)に対して書式設定を行う。
    3行目をヘッダーとして列名を判定する。
    指定列は左揃え、それ以外は右揃え + 3桁カンマ区切りにする。
    
    Args:
        ws: 対象のWorksheet
        left_align_cols: 左揃えにする列名のリスト
    """

    # 列インデックスと列名のマッピングを作成 (1始まり)
    left_col_indices = set()
    right_col_indices = set()

    # ヘッダー行のセルを読み込む
    for cell in ws[3]:
        # なぜか最初のシートだけNoneが4個認識されてしまうので弾く
        if not cell.value:
            break

        col_name = str(cell.value)

        # 指定された列名なら左揃えリストへ
        if col_name in left_align_cols:
            left_col_indices.add(cell.column)
        else:
            right_col_indices.add(cell.column)
        
    # 行ごとにスタイル適用(列で一気にやろうとしたらなぜか効かなかったため)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            if cell.column in left_col_indices:
                cell.alignment = Alignment(horizontal='left')
            elif cell.column in right_col_indices:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'


date_of_execution = datetime.now(ZoneInfo('Asia/Tokyo')).strftime('%Y/%m/%d')

def add_new_column_to_summarysheet_about_number_of_company(
    target_df: pd.DataFrame,
    book: Workbook,
    summary_sheetname: str,
) -> None:
    """【サマリ】1,2,3(企業)を作成"""

    try:
        # 参照渡しなので、これ以降summary_sheetを変更したらbookを変更したことにもなる
        summary_sheet = book[summary_sheetname]
        
        # 最後の列の番号を取得する (例: c列までデータがあれば、max_columnは3になる)
        last_col_number = summary_sheet.max_column
        # その次の列の番号(ここに今月分の値を代入していく)
        target_col_number = last_col_number + 1

        target_letter = get_column_letter(target_col_number)

        # 列の幅は25.75で固定（実際はなぜか25.17になる）
        summary_sheet.column_dimensions[target_letter].width = 25.75

        set_value_and_copy_style(summary_sheet, 4, target_col_number, date_of_execution)

        total_number_of_employees = target_df['社員数'].sum()
        set_value_and_copy_style(summary_sheet, 5, target_col_number, total_number_of_employees)

        if '3' in summary_sheetname:
            xxx = target_df['xxx'].sum()
            # ...
        else:
            yyy = target_df['yyy'].sum()
            # ...

        # ...

        # サマリシートの各テーブル範囲を1列増やす
        expand_table_range(summary_sheet)
    except Exception as e:
        logging.error(f"「{summary_sheetname}」シートへの書き込みもしくは値の計算に失敗しました: {e}")


def add_new_column_to_summarysheet_about_number_of_employee(
    target_df_user: pd.DataFrame | None,  # 3では使用しないためNoneで呼び出す
    target_df_office: pd.DataFrame,
    book: Workbook,
    summary_sheetname: str,
) -> None:
    """【サマリ】1,2,3(社員)を作成"""

    try:
        summary_sheet: Worksheet = book[summary_sheetname]

        last_col_number = summary_sheet.max_column
        target_col_number = last_col_number + 1

        target_letter = get_column_letter(target_col_number)
        summary_sheet.column_dimensions[target_letter].width = 25.75

        set_value_and_copy_style(summary_sheet, 4, target_col_number, date_of_execution)

        total_number_of_assigned_tasks = target_df_office['担当業務数'].mean()
        set_value_and_copy_style(summary_sheet, 5, target_col_number, total_number_of_assigned_tasks)

        if '3' in summary_sheetname:
            xxx = target_df_office['xxx'].sum()
            # ...
        else:
            yyy = target_df_user['yyy'].sum()
            # ...

        # ...

        expand_table_range(summary_sheet)
    except Exception as e:
        logging.error(f"「{summary_sheetname}」シートへの書き込みもしくは値の計算に失敗しました: {e}")


def set_value_and_copy_style(
    summary_sheet: Worksheet,
    row: int,
    col: int,
    value: int | float  # sum()の返り値はAnyという仕様だがint | float と宣言しておく
) -> None:
    """指定したセルに値を書き込み、すぐ左の列(col-1)のセルから書式(フォント、罫線、塗りつぶし、表示形式、配置)をコピー"""

    # セルへの値の書き込み(Setter的な使い方)
    cell = summary_sheet.cell(row=row, column=col, value=value)

    # セルの値の取得(Getter的な使い方。書き込みは行われない)
    source_cell = summary_sheet.cell(row=row, column=col - 1)
    
    if source_cell.has_style:
        cell.font = copy(source_cell.font)  # フォント
        cell.border = copy(source_cell.border)  # 白い罫線(グリッド線)
        cell.fill = copy(source_cell.fill)  # 塗りつぶし(背景色)
        cell.number_format = copy(source_cell.number_format)  # カンマ区切り
        cell.protection = copy(source_cell.protection)  # シート保護やセルのロック
        cell.alignment = copy(source_cell.alignment)  # 配置(右揃え など)


def expand_table_range(ws: Worksheet) -> None:
    """
    サマリシート内の各テーブル範囲(ref)を拡張し、不足している列定義(TableColumn)を追加する。
    TableColum追加まで行わないと「ファイルが破損しています」というエラーになる。
    """

    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)

        # 1. 範囲（ref）の更新
        new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col + 1)}{max_row}"
        table.ref = new_ref

        # 2. オートフィルタの範囲更新（設定されている場合）
        if table.autoFilter:
            table.autoFilter.ref = new_ref

        # 3. 列定義（TableColumn）の追加
        # idはユニークである必要がある
        current_id = len(table.tableColumns) + 1

        # ヘッダー行(min_row)の最終列(今回挿入した新しい列)のセルから値を取得して、列名にする。つまり最新の 年/月/日 になる
        # テーブルの列名は必須であり、かつ重複してはいけない
        # nameは日付型などではなく文字列型が必須のため str() で変換する
        header_val = ws.cell(row=min_row, column=max_col + 1).value
        str_header_val = str(header_val)
        
        # 定義を作成して追加
        new_col = TableColumn(id=current_id, name=str_header_val)
        table.tableColumns.append(new_col)


def create_replacements_dict(book: Workbook) -> dict[str, tuple[str, str]]:
    """replacements辞書を作成し返却"""

    summary_sheet_names = [
        '【サマリ】1(企業)', '【サマリ】1(社員)',
        '【サマリ】2(企業)', '【サマリ】2(社員)',
        '【サマリ】3(企業)', '【サマリ】3(社員)'
    ]
    replacements = {}

    for summary_sheet_name in summary_sheet_names:
        ws = book[summary_sheet_name]

        current_max_col = ws.max_column     # 現在の最終列（拡張後の列）
        prev_max_col = current_max_col - 1  # 拡張前の列（1つ左）
        
        old_letter = get_column_letter(prev_max_col)
        new_letter = get_column_letter(current_max_col)
        
        replacements[summary_sheet_name] = (old_letter, new_letter)
        logging.info(f"グラフのデータ範囲の最終列の置換ルール登録: {summary_sheet_name}シートの{old_letter}までを{new_letter}までに拡張")
    
    return replacements


def patch_xlsx_charts(
    input_stream: io.BytesIO,
    replacements: list[tuple[str, str]]
) -> io.BytesIO:
    """
    保存した後のxlsxファイル(zip)のバイナリ(ストリーム)を受け取り、内部のチャートXMLを直接書き換えて
    グラフのデータ範囲を拡張(1列分増やす)した新しいストリームを返す。
    サマリのテーブル拡張のようにopenpyxlでやろうとしたがグラフを認識できなかったためこの方法で行う。
    """
    
    # 読み取り位置を先頭に戻す
    input_stream.seek(0)
    
    # 出力用の新しいストリーム
    output_stream = io.BytesIO()
    
    # zipとして開き、ファイルをコピーしながら必要なら書き換え
    # 型1. ZipFile: Zipファイル全体
    # 型2. ZipInfo: ZipFileの中に入っている個々のファイルのメタ情報(filenameなど)
    with zipfile.ZipFile(input_stream, 'r') as zin:
        with zipfile.ZipFile(output_stream, 'w') as zout:
            # itemはzipfileの中の各xml
            for item in zin.infolist():
                data = zin.read(item.filename)

                # チャート定義ファイルの場合のみ置換処理を行う
                if item.filename.startswith('xl/charts/chart') and item.filename.endswith('.xml'):
                    # XMLはバイト列なので文字列にデコード
                    xml_str = data.decode('utf-8')

                    for summary_sheet_name in replacements:
                        # 見つけた.xmlがどのサマリシートを参照しているのか、特定するまでループ
                        # 特定したらそのサマリの拡張前・拡張後の列名のペアを取り出し、それを.xmlに適用(1列拡張)する
                        if summary_sheet_name in xml_str:
                            old_col, new_col = replacements[summary_sheet_name]
                            # 正規表現: コロン(:) + $ + 旧列文字 + $ + 数字
                            # 例: E列をF列にする場合、 :$E$5 -> :$F$5 に置換する
                            # これにより範囲の「終了位置」だけが伸びる
                            pattern = f"(:\\$){old_col}(\\$\\d+)"
                            repl = f"\\g<1>{new_col}\\g<2>"
                            
                            # re.sub(正規表現, 正規表現にマッチした部分の置換後の文字列, 置換対象の文字列)
                            xml_str = re.sub(pattern, repl, xml_str)
                            break
                    
                    # 書き換えたデータをUTF-8バイト列に戻す
                    data = xml_str.encode('utf-8')
                
                # 新しいzipに書き込み
                zout.writestr(item, data)
    
    # ポインタを先頭に戻して返す
    output_stream.seek(0)

    return output_stream


# Azure Blob Storageのコンテナ内にある、各SQLファイルのURL
query_url_per_employee_1 = get_env_or_raise('QUERY_URL_PER_employee_1')
query_url_per_company_1 = get_env_or_raise('QUERY_URL_PER_company_1')
query_url_per_employee_2 = get_env_or_raise('QUERY_URL_PER_employee_2')
query_url_per_company_2 = get_env_or_raise('QUERY_URL_PER_company_2')
query_url_per_employee_3 = get_env_or_raise('QUERY_URL_PER_employee_3')
query_url_per_company_3 = get_env_or_raise('QUERY_URL_PER_company_3')

query_url_get_employeename_1 = get_env_or_raise('QUERY_URL_GET_employeeNAME_1')
query_url_get_companyname_1 = get_env_or_raise('QUERY_URL_GET_companyNAME_1')
query_url_get_employeename_2 = get_env_or_raise('QUERY_URL_GET_employeeNAME_2')
query_url_get_companyname_2 = get_env_or_raise('QUERY_URL_GET_companyNAME_2')

# SQLファイルのURLと、それに対応するベース名(シート名の一部として後に使用する)
SQL_FILES = [
    (query_url_per_employee_1, "1(企業ごと)"),
    (query_url_per_company_1, "1(社員ごと)"),
    (query_url_per_employee_2, "2(企業ごと)"),
    (query_url_per_company_2, "2(社員ごと)"),
    (query_url_per_employee_3, "3(企業ごと)"),
    (query_url_per_company_3, "3(社員ごと)"),
]

GETNAME_SQL_FILES = [
    (query_url_get_employeename_1, "1(企業名取得)"),
    (query_url_get_companyname_1, "1(社員名取得)"),
    (query_url_get_employeename_2, "2(企業名取得)"),
    (query_url_get_companyname_2, "2(社員名取得)"),
]


def main_process():
    # ---環境変数を取得---
    driver = get_env_or_raise('AZURE_SQL_DRIVER')
    
    # 1
    server_1 = get_env_or_raise('SERVER_1')
    database_1 = get_env_or_raise('DATABASE_1')
    username_1 = get_env_or_raise('USERNAME_1')
    password_1 = get_env_or_raise('PASSWORD_1')

    conn_str_1 = (
        f'DRIVER={driver};'
        f'SERVER={server_1};'
        f'DATABASE={database_1};'
        f'UID={username_1};'
        f'PWD={password_1};'
        'Encrypt=yes;'
        'TrustServerCertificate=no;'
        'Connection Timeout=120;'  # Azure SQL Databaseへの接続確立の制限時間
    )

    # 2
    server_2 = get_env_or_raise('SERVER_2')
    database_2 = get_env_or_raise('DATABASE_2')
    username_2 = get_env_or_raise('USERNAME_2')
    password_2 = get_env_or_raise('PASSWORD_2')

    conn_str_2 = (
        f'DRIVER={driver};'
        f'SERVER={server_2};'
        f'DATABASE={database_2};'
        f'UID={username_2};'
        f'PWD={password_2};'
        'Encrypt=yes;'
        'TrustServerCertificate=no;'
        'Connection Timeout=120;'
    )

    # 3
    server_3 = get_env_or_raise('SERVER_3')
    database_3 = get_env_or_raise('DATABASE_3')
    username_3 = get_env_or_raise('USERNAME_3')
    password_3 = get_env_or_raise('PASSWORD_3')

    conn_str_3 = (
        f'DRIVER={driver};'
        f'SERVER={server_3};'
        f'DATABASE={database_3};'
        f'UID={username_3};'
        f'PWD={password_3};'
        'Encrypt=yes;'
        'TrustServerCertificate=no;'
        'Connection Timeout=120;'
    )

    # 1で、企業名/社員名がNULLだった場合に利用する接続情報
    getname_server_1 = get_env_or_raise('GETNAME_SERVER_1')
    getname_database_1 = get_env_or_raise('GETNAME_DATABASE_1')
    getname_username_1 = get_env_or_raise('GETNAME_USERNAME_1')
    getname_password_1 = get_env_or_raise('GETNAME_PASSWORD_1')

    getname_conn_str_1 = (
        f'DRIVER={driver};'
        f'SERVER={getname_server_1};'
        f'DATABASE={getname_database_1};'
        f'UID={getname_username_1};'
        f'PWD={getname_password_1};'
        'Encrypt=yes;'
        'TrustServerCertificate=no;'
        'Connection Timeout=120;'
    )

    # 2で、企業名/社員名がNULLだった場合に利用する接続情報
    getname_server_2 = get_env_or_raise('GETNAME_SERVER_2')
    getname_database_2 = get_env_or_raise('GETNAME_DATABASE_2')
    getname_username_2 = get_env_or_raise('GETNAME_USERNAME_2')
    getname_password_2 = get_env_or_raise('GETNAME_PASSWORD_2')

    getname_conn_str_2 = (
        f'DRIVER={driver};'
        f'SERVER={getname_server_2};'
        f'DATABASE={getname_database_2};'
        f'UID={getname_username_2};'
        f'PWD={getname_password_2};'
        'Encrypt=yes;'
        'TrustServerCertificate=no;'
        'Connection Timeout=120;'
    )

    # 3ではNULLにならないため省略

    # シート名を格納
    basename_df_list: list[tuple[str, pd.DataFrame]] = []

    # データ取得のループ
    for sql_url, base_name in SQL_FILES:
        logging.info(f"SQL_FILESループ -> {sql_url}")
        
        try:
            sql_text = download_sql_text(sql_url)
        except Exception as e:
            logging.error(f"エラー: download_sql_text()に失敗しました: {sql_url}")
            logging.error(f"エラー内容: {e}")
            continue

        if '1' in base_name:
            conn_str = conn_str_1
        elif '2' in base_name:
            conn_str = conn_str_2
        else:
            conn_str = conn_str_3

        try:
            df = execute_sql_to_df(conn_str, sql_text)
        except Exception as e:
            logging.error(f"エラー! execute_sql_to_df()に失敗しました: {base_name}")
            logging.error(f"エラー内容: {e}")
            basename_df_list.append((f"エラー_{base_name}", None))
            continue

        basename_df_list.append((base_name, df))
        logging.info('現在のループ内の処理正常終了')

    # 名前情報を格納する4つのdfを格納するリスト
    """
    ループの流れ
    1回目: 企業コード(1), 企業名(1)
    2回目: 社員コード(1), 社員名(1)
    3回目: 企業コード(2), 企業名(2)
    4回目: 社員コード(2), 社員名(2)
    """
    names_df_list: list[pd.DataFrame] = []

    # 名前取得のループ
    for sql_url, base_name in GETNAME_SQL_FILES:
        logging.info(f"GETNAME_SQL_FILESループ -> {sql_url}")

        # 実行したいsql文をダウンロード
        try:
            sql_text = download_sql_text(sql_url)
        except Exception as e:
            logging.error(f"エラー! download_sql_text()に失敗しました: {sql_url} : {e}")

        if '1' in base_name:
            conn_str = getname_conn_str_1
        elif '2' in base_name:
            conn_str = getname_conn_str_2

        # dfのNULLになっている名前を置き換えるための名前情報が含まれたdfを作成
        try:
            names_df = execute_sql_to_df(conn_str, sql_text)
            names_df_list.append(names_df)
        except Exception as e:
            logging.error(f"エラー! execute_sql_to_df()に失敗しました: {sql_url}")
            logging.error(f"エラー内容: {e}")

        # dfのNULLをnames_dfによって置き換え
        try:
            # 1回目のループ: 1(企業名取得)
            if sql_url == GETNAME_SQL_FILES[0][0]:
                df = basename_df_list[0][1]
                replace_null(df, names_df_list, 0, '企業コード', '企業名')
        except Exception as e:
            logging.error(f"エラー! 1巡目のreplace_null()に失敗しました: {sql_url}")
            logging.error(f"エラー内容: {e}")

        try:
            # 2回目のループ: 1(企業名&社員名取得)
            if sql_url == GETNAME_SQL_FILES[1][0]:
                df = basename_df_list[1][1]
                replace_null(df, names_df_list, 0, '企業コード', '企業名')
                replace_null(df, names_df_list, 1, '社員コード', '社員名')
        except Exception as e:
            logging.error(f"エラー! 2巡目のreplace_null()に失敗しました: {sql_url}")
            logging.error(f"エラー内容: {e}")

        try:
            # 3回目のループ: 2(企業名取得)
            if sql_url == GETNAME_SQL_FILES[2][0]:
                df = basename_df_list[2][1]
                replace_null(df, names_df_list, 2, '企業コード', '企業名')
        except Exception as e:
            logging.error(f"エラー! 3巡目のreplace_null()に失敗しました: {sql_url}")
            logging.error(f"エラー内容: {e}")

        try:
            # 4回目のループ: 2(企業名&社員名取得)
            if sql_url == GETNAME_SQL_FILES[3][0]:
                df = basename_df_list[3][1]
                replace_null(df, names_df_list, 2, '企業コード', '企業名')
                replace_null(df, names_df_list, 3, '社員コード', '社員名')
        except Exception as e:
            logging.error(f"エラー! 4巡目のreplace_null()に失敗しました: {sql_url}")
            logging.error(f"エラー内容: {e}")

        logging.info('現在のループ内の処理終了')

    # azureとの接続関連
    CONNECTION_STRING = get_env_or_raise('CONNECTION_STRING')
    CONTAINER_NAME = get_env_or_raise('CONTAINER_NAME')
    EXCEL_FILE_NAME_PREFIX = get_env_or_raise('EXCEL_FILE_NAME_PREFIX')

    # クライアントの初期化
    blob_service_client = BlobServiceClient.from_connection_string(
        CONNECTION_STRING,
        connection_timeout=600,  # 接続確立までの待機秒数
        read_timeout=600,  # データ読み込みの待機秒数
        retry_total=5  # 失敗時のリトライ回数
    )
    container_client = blob_service_client.get_container_client(CONTAINER_NAME)

    # 最新(先月20日時点)のエクセルを見つける
    # それより前の月のエクセルはファイル名に_oldをつけてアーカイブ扱いしているので、_oldがないものが最新ということになる
    blobs = list(container_client.list_blobs(name_starts_with=EXCEL_FILE_NAME_PREFIX))

    # listに格納されたエクセルの中から最新のものを見つける
    for blob in blobs:
        if '_old' not in blob.name:
            latest_excel_blob_name = blob.name
            break

    # Blobからエクセルファイルを仮想メモリにダウンロード
    # pd.read_excel()だけで済ませてしまうと、エクセルとして持ってくるのではなく中身のデータしか持ってこないからデザインが消えたり色々な問題があるのでバイトで扱う。その後にピンポイントでデータを置き換えたいところでだけpandas使用していく
    # get_blob_clientはgetというよりcreateが実態に近い
    blob_client = container_client.get_blob_client(latest_excel_blob_name)

    # メモリ上に空の仮想ファイルを作成
    # コンピュータのディスク(HDD/SSD)ではなくメモリ(RAM)に作成される
    download_stream = io.BytesIO()

    # クラウドからデータをダウンロードし、仮想ファイルに流し込む(書き込む)
    blob_client.download_blob().readinto(download_stream)

    # 読み取り位置(カーソル)が最後になっているので、先頭(0バイト目)に戻す
    # これをやらないと pd.read_excel(download_stream) などやってもデータが空と判断されてエラーになる
    download_stream.seek(0)

    # excelの基になるbookを作成
    book = load_workbook(download_stream)
    output_stream = io.BytesIO()

    # ---一覧表(6シート)の作成---
    # 今の年月(新規作成するシート名に使用)
    now_year_dot_month = datetime.now(ZoneInfo('Asia/Tokyo')).strftime('%Y.%m')
    now = datetime.now(ZoneInfo('Asia/Tokyo'))

    # 1か月前の年月(削除するシート名の判別に使用)
    prev_date = pd.Timestamp(datetime.now(ZoneInfo('Asia/Tokyo'))) - pd.DateOffset(months=1)
    prev_year_dot_month = prev_date.strftime('%Y.%m')

    for basename, df in basename_df_list:
        new_sheetname = f"【{now_year_dot_month}】{basename}"
        old_sheetname = f"【{prev_year_dot_month}】{basename}"

        ws = book[old_sheetname]

        # シート名を変更(年月日を今月のものに変える)
        ws.title = new_sheetname

        # 既存のシートのデータを削除(ヘッダーである3行目まではそのままでいいので除く)
        # 最大行数(max_row)分消してしまうと、100行のデータがあったら100行目まで消せば良いはずなのに4~103行目まで消されてしまうので無駄な処理が入る
        ws.delete_rows(4, amount=ws.max_row - 3)

        # データの書き込み
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

        match basename:
            case '1(企業ごと)':
                # A1セルの値の年月日を現在のものに変える
                ws['A1'].value = f'1(企業ごと)利用集計({now.year}年{now.month}月{now.day}日時点累計数)'
                # 書式設定(左揃え、右揃え、桁区切り)の適用
                logging.info(f'{basename}のapply_column_style()を実行中...')
                apply_column_style(ws, ['企業コード', '企業名'])
            case '1(社員ごと)':
                ws['A1'].value = f'1(社員ごと)利用集計({now.year}年{now.month}月{now.day}日時点累計数)'
                logging.info(f'{basename}のapply_column_style()を実行中...')
                apply_column_style(ws, ['企業コード', '企業名', '社員コード', '社員名'])
            case '2(企業ごと)':
                ws['A1'].value = f'2(企業ごと)利用集計({now.year}年{now.month}月{now.day}日時点累計数)'
                logging.info(f'{basename}のapply_column_style()を実行中...')
                apply_column_style(ws, ['企業コード', '企業名'])
            case '2(社員ごと)':
                ws['A1'].value = f'2(社員ごと)利用集計({now.year}年{now.month}月{now.day}日時点累計数)'
                logging.info(f'{basename}のapply_column_style()を実行中...')
                apply_column_style(ws, ['企業コード', '企業名', '社員コード', '社員名'])
            case '3(企業ごと)':
                # 2列しかなくスタイル適用処理が必要ない
                ws['A1'].value = f'3(企業ごと)利用集計({now.year}年{now.month}月{now.day}日時点累計数)'
            case '3(社員ごと)':
                ws['A1'].value = f'3(社員ごと)利用集計({now.year}年{now.month}月{now.day}日時点累計数)'
                logging.info(f'{basename}のapply_column_style()を実行中...')
                # 企業名は取得する必要なし（集計の都合上）
                apply_column_style(ws, ['企業コード', '企業コード2', '社員コード', '社員名'])


    # ---サマリ(6シート)の作成---
    logging.info('「【サマリ】1(企業)」に1列追加中...')
    add_new_column_to_summarysheet_about_number_of_company(
        basename_df_list[1][1],
        book,
        '【サマリ】1(企業)',
    )

    logging.info('「【サマリ】1(社員)」に1列追加中...')
    add_new_column_to_summarysheet_about_number_of_employee(
        basename_df_list[0][1],
        basename_df_list[1][1],
        book,
        '【サマリ】1(社員)',
    )

    logging.info('「【サマリ】2(企業)」に1列追加中...')
    add_new_column_to_summarysheet_about_number_of_company(
        basename_df_list[3][1],
        book,
        '【サマリ】2(企業)',
    )

    logging.info('「【サマリ】2(社員)」に1列追加中...')
    add_new_column_to_summarysheet_about_number_of_employee(
        basename_df_list[2][1],
        basename_df_list[3][1],
        book,
        '【サマリ】2(社員)',
    )

    logging.info('「【サマリ】3(企業)」に1列追加中...')
    add_new_column_to_summarysheet_about_number_of_company(
        basename_df_list[5][1],
        book,
        '【サマリ】3(企業)'
    )

    logging.info('「【サマリ】3(社員)」に1列追加中...')
    add_new_column_to_summarysheet_about_number_of_employee(
        None,
        basename_df_list[5][1],
        book,
        '【サマリ】3(社員)'
    )

    # ---グラフのデータ範囲を1列分拡張(6シート)---

    # 編集が終わったbookをoutput_streamに保存
    logging.info('編集が終わったbookをoutput_streamに保存中...')
    book.save(output_stream)

    # シート名(文字列)をキーとして、拡張前の列名と拡張後の列名が入ったタプルを格納する辞書を作成
    replacements = create_replacements_dict(book)

    # 関数を呼んでストリームの中身を書き換える
    logging.info("グラフ範囲のXML直接置換を実行中...")
    output_stream = patch_xlsx_charts(output_stream, replacements)

    # 新しいファイル名でアップロード(実行時の年月日を使用)
    today_str = datetime.now(ZoneInfo('Asia/Tokyo')).strftime('%Y%m%d')
    new_excel_blob_name = f'{EXCEL_FILE_NAME_PREFIX}{today_str}.xlsx'

    logging.info('get_blob_client(new_excel_blob_name)開始...')
    new_blob_client = container_client.get_blob_client(new_excel_blob_name)

    # timeout: 処理全体のタイムアウト秒数
    # max_concurrency: 並列アップロード数（デフォルトは1。増やすと速くなるが、不安定な回線では1か2が良い）
    logging.info('upload_blob()開始...')
    new_blob_client.upload_blob(output_stream, overwrite=True, timeout=600, max_concurrency=2)

    logging.info(f'新規ファイルをアップロードしました: {new_excel_blob_name}')

    # 古いファイルの名前に _old をつける(Azure Blobにはリネームのコマンドがないため、コピーして削除する)
    old_renamed_blob_name = latest_excel_blob_name.replace('.xlsx', '_old.xlsx')

    logging.info('get_blob_client(old_renamed_blob_name)開始...')
    old_blob_client = container_client.get_blob_client(old_renamed_blob_name)

    # 先月時点のblobのコピーとして _old というsuffixをつけたblobをコピーによりコンテナ上で作成
    logging.info('start_copy_from_url()開始...')
    old_blob_client.start_copy_from_url(blob_client.url)

    # 元のファイルを削除
    logging.info('delete_blob()開始...')
    blob_client.delete_blob()

    logging.info(f'古いファイルをリネームしました: {old_renamed_blob_name}')


# ローカルで動かす時
# if __name__ == '__main__':
#     try:
#         main_process()
#         logging.info('処理がすべて終了しました')
#     except Exception as e:
#         logging.info(f"main_process実行中にエラー発生: {e}")
#         raise

        
# schedule: "秒 分 時 日 月 曜日"
# CRON式: https://learn.microsoft.com/ja-jp/azure/azure-functions/functions-bindings-timer?tabs=python-v2%2Cisolated-process%2Cnodejs-v4&pivots=programming-language-python#ncrontab-expressions
# Asia/Tokyoで毎月20日の4:00に実行したい -> UTCだと9時間前なので19日の19:00になる
@app.schedule(
    schedule="0 0 19 19 * *",
    # schedule="0 0 */2 * * *",  # デバッグ用: 2時間おきに実行
    # schedule="0 0 * * * *",  # デバッグ用: 1時間おきに実行
    arg_name="myTimer",
    run_on_startup=False,
    use_monitor=False
) 
def monthly_processing(myTimer: func.TimerRequest) -> None:
    logging.info('Python timer trigger function started.')
    
    try:
        main_process()
        logging.info('処理が正常に終了しました。')
    except Exception as e:
        logging.error(f'処理中にエラーが発生しました: {e}')
        raise  # エラーを再送出してAzure側で失敗として記録させる
